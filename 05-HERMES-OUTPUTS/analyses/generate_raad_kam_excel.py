import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Raad KAM Merchants"

# Header
headers = ["#", "Merchant ID", "Merchant Name", "Phone"]
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Data
merchants = [
    (1, 6798, "Vapor World", "01823211114"),
    (2, 77290, "Startech & Engineering Ltd", "01313717006"),
    (3, 59693, "Nagad", "01704161133"),
    (4, 119653, "Arogga LTD", "01332537443"),
    (5, 147238, "Ranks Petroleum Limited_BD", "01730085007"),
    (6, 209286, "Daraz Bangladesh Limited", "01711111100"),
    (7, 234316, "Cartup", "01335088333"),
    (8, 302566, "Daraz Bangladesh Limited (CP)", "01911111100"),
    (9, 374756, "Cartup (Closed Box)", "01335088328"),
    (10, 384721, "Redx Open Box", "01676914776"),
    (11, 384722, "Redx Closed Box", "01313718045"),
]

data_font = Font(size=11)
data_align = Alignment(vertical="center")

for row_num, (sn, mid, name, phone) in enumerate(merchants, 2):
    ws.cell(row=row_num, column=1, value=sn).border = thin_border
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="center", vertical="center")
    
    cell_id = ws.cell(row=row_num, column=2, value=mid)
    cell_id.border = thin_border
    cell_id.alignment = Alignment(horizontal="center", vertical="center")
    cell_id.number_format = '0'
    
    cell_name = ws.cell(row=row_num, column=3, value=name)
    cell_name.border = thin_border
    cell_name.alignment = data_align
    
    cell_phone = ws.cell(row=row_num, column=4, value=phone)
    cell_phone.border = thin_border
    cell_phone.alignment = Alignment(horizontal="center", vertical="center")

# Column widths
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 16

# Info row
row = len(merchants) + 3
ws.cell(row=row, column=1, value="KAM:").font = Font(bold=True)
ws.cell(row=row, column=2, value="Ahnaf Tahmid Raad (Key Account Management)").font = Font(italic=True)
ws.cell(row=row+1, column=1, value="Source:").font = Font(bold=True)
ws.cell(row=row+1, column=2, value="public_merchants.kam_id = 13507").font = Font(italic=True, color="666666")
ws.cell(row=row+2, column=1, value="Generated:").font = Font(bold=True)
from datetime import datetime
ws.cell(row=row+2, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M BDT")).font = Font(italic=True, color="666666")

output_path = "/home/ubuntu/Hermes_Knowledge_Base/05-HERMES-OUTPUTS/analyses/raad_kam_merchants.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
